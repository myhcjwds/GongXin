# word -> excel 后，  缺少 企业名称 和 纳税人识别号
# 需要从陈泽给的excel中查询  ，并且填补到 上面的excel中
import os

from openpyxl import load_workbook
from WordToExcel.path_tool import get_output_folder_path

# 包含待填补信息的查询表
select_xlsx_path = r'缺失信息查询表.xlsx'
workbook2 = load_workbook(select_xlsx_path)
sheet2 = workbook2.active


# 需要添加列的excel所在文件夹
origin_excel_folder_path = get_output_folder_path()


# 遍历文件夹中的所有.xls文件
for filename in os.listdir(origin_excel_folder_path):
    if filename.endswith('.xlsx'):


        xlsx_path = os.path.join(origin_excel_folder_path, filename)

        workbook = load_workbook(xlsx_path)

        # 获取指定的Sheet
        sheet = workbook.active

        # 要插入的列索引
        insert_column_index_of_mingcheng = 3
        # 要插入的列名称
        insert_column_value_of_mingcheng = "企业名称"
        sheet.insert_cols(insert_column_index_of_mingcheng)
        sheet.cell(row=1, column=3).value = insert_column_value_of_mingcheng

        # 要插入的列索引
        insert_column_index_of_shuihao = 5
        # 要插入的列名称
        insert_column_value_of_shuihao  = "税号"
        sheet.insert_cols(insert_column_index_of_shuihao)
        sheet.cell(row=1, column=5).value = insert_column_value_of_shuihao

        # 设置新插入列的值（假设我们为第2行到第5行设置值）
        for i in range(2, sheet.max_row + 1):
            # 插入对应行的企业名称 和 税号
            for j in range(2, sheet2.max_row + 1):
                # 先根据 信用代码 找到 匹配的行
                if sheet.cell(row=i, column=4).value ==  sheet2.cell(row=j, column=20).value:

                    # 插入对应行的企业名称
                    sheet.cell(row=i, column=insert_column_index_of_mingcheng).value = sheet2.cell(row=j, column=1).value
                    # 插入对应行的税号
                    sheet.cell(row=i, column=insert_column_index_of_shuihao).value = sheet2.cell(row=j, column=21).value

                    break




        # 保存每个excel的更新
        workbook.save(xlsx_path)

